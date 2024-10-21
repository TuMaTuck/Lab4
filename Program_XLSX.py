import csv
import openpyxl
from datetime import datetime

def calculate_age(birth_date):
    today = datetime.today()
    return today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

# Читання CSV файлу
employees = []
try:
    with open('employees.csv', mode='r', encoding='utf-8') as file:
        reader = csv.reader(file)
        next(reader)  # Пропускаємо заголовок
        for row in reader:
            birth_date = datetime.strptime(row[4], '%Y-%m-%d')
            age = calculate_age(birth_date)
            employees.append([*row[:4], birth_date.date(), age])
except FileNotFoundError:
    print("Файл CSV не знайдено!")
    exit()

# Створення Excel файлу
wb = openpyxl.Workbook()
sheets = {
    "all": wb.active,
    "younger_18": wb.create_sheet("younger_18"),
    "18-45": wb.create_sheet("18-45"),
    "45-70": wb.create_sheet("45-70"),
    "older_70": wb.create_sheet("older_70"),
}

headers = ['№', 'Прізвище', 'Ім’я', 'По батькові', 'Дата народження', 'Вік']

for sheet in sheets.values():
    sheet.append(headers)

# Розподіл по аркушах
for i, employee in enumerate(employees, 1):
    sheets["all"].append([i] + employee)
    age = employee[5]

    if age < 18:
        sheets["younger_18"].append([i] + employee)
    elif 18 <= age <= 45:
        sheets["18-45"].append([i] + employee)
    elif 45 < age <= 70:
        sheets["45-70"].append([i] + employee)
    else:
        sheets["older_70"].append([i] + employee)

# Збереження Excel файлу
try:
    wb.save('employees.xlsx')
    print("Ok")
except Exception as e:
    print("Помилка при створенні XLSX файлу:", e)
